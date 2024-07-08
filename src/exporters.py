import jsonlines
from src.logger import logger
import os
from typing import Dict, List
import csv
import json
from openpyxl import Workbook, load_workbook



class JsonExporter:

    def __init__(self, filename: str):
        self.filename = filename
        self.mode = 'a' if os.path.exists(filename) else 'w'
        self.file = open(self.filename, mode=self.mode)

    def export(self, items: List[Dict]):
        if not items:
            logger.info(f"No data available")
            return
        try:
            json.dump(items, self.file)
        except Exception as e:
            logger.debug(f"Error while writing to {self.filename}")
        else:
            logger.info(f"Data written to {self.filename}")

    def close(self):
        try:
            self.file.close()
        except IndexError as e:
            logger.debug(f"Error while closing the {self.filename}: {e}")



class JsonlExporter:

    def __init__(self, filename: str):
        self.filename = filename
        self.mode = 'a' if os.path.exists(filename) else 'w'
        self.file = jsonlines.open(self.filename, mode=self.mode)

    def export(self, items: List[Dict]):
        if not items:
            logger.info(f"No data available")
            return
        try:
            self.file.write_all(items)
        except Exception as e:
            logger.debug(f"Error while writing to {self.filename}")
        else:
            logger.info(f"Data written to {self.filename}")

    def close(self):
        try:
            self.file.close()
        except IndexError as e:
            logger.debug(f"Error while closing the {self.filename}: {e}")


class CsvExporter:

    def __init__(self, filename: str, fields: List):
        self.filename = filename
        self.mode = 'a' if os.path.exists(filename) else 'w'
        self.file = open(self.filename, mode=self.mode)
        self.writer = csv.DictWriter(self.file, fieldnames=fields)
        self.writer.writeheader()

    def export(self, items: List[Dict]):
        if not items:
            logger.info(f"No data available")
            return
        try:
            self.writer.writerows(items)
        except Exception as e:
            logger.debug(f"Error while writing to {self.filename}")
        else:
            logger.info(f"Data written to {self.filename}")

    def close(self):
        try:
            self.file.close()
        except IndexError as e:
            logger.debug(f"Error while closing the {self.filename}: {e}")



class XlsxExporter:

    def __init__(self, filename: str):
        self.filename = filename
        if os.path.exists(filename):
            self.workbook = load_workbook(filename)
        else:
            self.workbook = Workbook()
            # Remove the default sheet created by Workbook()
            self.workbook.remove(self.workbook.active)

    def export(self, items: List[Dict], sheet: str):
        if not items:
            logger.info(f"No data available for {sheet}")
            return
        try:
            if sheet in self.workbook.sheetnames:
                ws = self.workbook[sheet]
                self.workbook.remove(ws)
            ws = self.workbook.create_sheet(sheet)
            headers = list(items[0].keys())
            ws.append(headers)
            for item in items:
                ws.append([item.get(header, '') for header in headers])
        except Exception as e:
            logger.debug(f"Error while writing to {sheet}: {e}")
        else:
            logger.info(f"Data written to {sheet}")

    def close(self):
        try:
            self.workbook.save(self.filename)
        except Exception as e:
            logger.debug(f"Error while saving workbook: {e}")
        finally:
            self.workbook.close()